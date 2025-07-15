import z_add_path  # noqa: F401
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from cobrak.io import ensure_folder_existence, get_files, json_load, standardize_folder
from cobrak.spreadsheet_functionality import BG_COLOR_GREEN, BG_COLOR_RED
from cobrak.utilities import sort_dict_keys
from cobrak.constants import OBJECTIVE_VAR_NAME


BG_COLOR_GREY = PatternFill(start_color="d3d3d3", end_color="d3d3d3", fill_type="solid")


def _choose_color(current_value: float, wt_value: float):
    if max(wt_value / current_value, current_value / wt_value) < 1.001:
        return BG_COLOR_GREY
    if wt_value > current_value:
        return BG_COLOR_RED
    return BG_COLOR_GREEN


minmu_ac_ectfva_data = json_load(
    "./examples/iCH360/prepared_external_resources/variability_dict_enforced_growth_and_ac.json"
)
cobrak_model = json_load(
    "./examples/iCH360/prepared_external_resources/iCH360_cobrak.json"
)
raw_ko_targets = [
    reac_id
    for reac_id, (min_flux, max_flux) in minmu_ac_ectfva_data.items()
    if (min_flux <= 0.0)
    and (max_flux != 0.0)
    and (reac_id in cobrak_model["reactions"])
    and ("EX_" not in reac_id)
]
found_ko_targets = []
ko_targets = []
for raw_ko_target in raw_ko_targets:
    if raw_ko_target in found_ko_targets:
        continue

    if raw_ko_target.endswith("_fw"):
        tried_other_id = (raw_ko_target + "\b").replace("_fw\b", "_bw")
    else:
        tried_other_id = (raw_ko_target + "\b").replace("_bw\b", "_fw")
    ko_target_1 = raw_ko_target

    if (tried_other_id in cobrak_model["reactions"]) and (
        tried_other_id not in found_ko_targets
    ):
        ko_target_2 = tried_other_id
    else:
        ko_target_2 = ""
    ko_targets.append((ko_target_1, ko_target_2))
    found_ko_targets.extend((ko_target_1, ko_target_2))
found_ko_targets = [x for x in found_ko_targets if x != ""]  # noqa: PLC1901


biomass_reac_id: str = "Biomass_fw"
acetate_reac_id: str = "EX_ac_e_fw"
glucose_reac_id: str = "EX_glc__D_e_bw"
datafolder: str = standardize_folder("examples/iCH360/RESULTS_SINGLES_KOS")
eligible_jsons: list[str] = [
    filename
    for filename in get_files(datafolder)
    if ("_time" not in filename)
    and filename.endswith(".json")
    and ("final_best_result_" in filename)
    and ("__ko" in filename)
]
print(eligible_jsons)

checkdict = {}
for eligible_json in eligible_jsons:
    number = int(eligible_json.split("__")[1].split("__")[0])
    kotarget = eligible_json.split("__")[2].replace(".json", "")
    print(number, kotarget)
    if kotarget not in checkdict:
        checkdict[kotarget] = [None, None, None]
    data = json_load("examples/iCH360/RESULTS_SINGLES_KOS/" + eligible_json)
    checkdict[kotarget][number - 1] = [
        round(data[OBJECTIVE_VAR_NAME], 4),
        round(data[biomass_reac_id], 4),
        round(data[acetate_reac_id], 4),
    ]

kotarget_to_max_obj = {}
for key, value in checkdict.items():
    for index in (0, 1, 2):
        try:
            assert value[0][index] == value[1][index] and value[0][index] == value[2][index]
        except AssertionError:
            print(key, value, index)
            print(value[0][index], value[1][index], value[2][index])
            print("---")
    objvalues = [value[0][0], value[0][1], value[0][2]]
    kotarget_to_max_obj[key] = objvalues.index(max(objvalues))
print(kotarget_to_max_obj)
"""
needed_prefixes = [
    "final_best_result__1__ko",
    "final_best_result__2__ko",
    "final_best_result__3__ko",
]
for ko_target in ko_targets:
    for needed_prefix in needed_prefixes:
        found = False
        one = needed_prefix + f"{ko_target[0]}_{ko_target[1]}" + ".json"
        two = needed_prefix + f"{ko_target[0]}_{ko_target[1]}" + "_.json"
        if (one in eligible_jsons) or (two in eligible_jsons):
            found = True
        if not found:
            print(
                "(",
                ko_target,
                ",",
                needed_prefix.split("__")[1].replace("__", ""),
                "),",
            )
"""

results: dict[str, dict[str, float]] = {
    "(Wild-type)": json_load(
        "examples/iCH360/RESULTS_GLCUPTAKE/final_best_result__1_maxglc1000.json"
    )
}
further_results: dict[str, dict[str, float]] = sort_dict_keys(
    {
        eligible_json.split("__ko")[1].replace(".json", ""): json_load(
            datafolder + eligible_json
        )
        for eligible_json in eligible_jsons
    }
)
further_results = {
    key[2:]: json_load(
        f"examples/iCH360/RESULTS_SINGLES_KOS/final_best_result__{kotarget_to_max_obj[key]}__{key}.json"
    ) for key in kotarget_to_max_obj
}
for key, value in further_results.items():
    results[key] = value

# Create a new workbook
workbook = Workbook()

# Get the active worksheet
worksheet = workbook.active

# Set the title of the worksheet
worksheet.title = "Acetate KOs result"

# Add some data to the worksheet
worksheet["A1"] = "Supplementary File 2"
worksheet["A2"] = (
    "COBRA-k: a powerful framework bridging constraint-based and kinetic metabolic modeling"
)
worksheet["A3"] = "Pavlos Stephanos Bekiaris¹, Steffen Klamt¹﹡"
worksheet["A4"] = (
    "¹Max Planck Institute for Dynamics of Complex Technical Systems, Magdeburg, Sandtorstr. 1"
)
worksheet["A5"] = "*Corresponding author: klamt@mpi-magdeburg.mpg.de"
worksheet["A6"] = ""
worksheet["A7"] = (
    "Table S2. Single Knockout analysis result fluxes and yields. Colors indicate values equal (grey), higher (green) and lower (red) than in the wild-type solution."
)
worksheet["A8"] = "KO target reaction ID"
worksheet["B8"] = "Growth µ [h⁻¹]"
worksheet["C8"] = "Glucose uptake [mmol⋅gDW⁻¹⋅h⁻¹]"
worksheet["D8"] = "Acetate excretion [mmol⋅gDW⁻¹⋅h⁻¹]"
worksheet["E8"] = "Ac/µ yield [mmol⋅gDW⁻¹]"
worksheet["F8"] = "Ac/Glc yield [-]"

worksheet.cell(row=1, column=1).font = Font(bold=True)
worksheet.cell(row=7, column=1).font = Font(italic=True)
for i in range(6):
    worksheet.cell(row=8, column=i + 1).font = Font(bold=True)

wt_result = results["(Wild-type)"]
wt_biomass_flux, wt_glucose_flux, wt_acetate_flux = (
    wt_result[biomass_reac_id],
    wt_result[glucose_reac_id],
    wt_result[acetate_reac_id],
)
wt_biomass_yield, wt_substrate_yield = (
    wt_acetate_flux / wt_biomass_flux,
    wt_acetate_flux / wt_glucose_flux,
)

start_line = 9
for current_line, (ko_target, result) in enumerate(results.items()):
    if not ko_target:
        continue
    biomass_flux, glucose_flux, acetate_flux = (
        result[biomass_reac_id],
        result[glucose_reac_id],
        result[acetate_reac_id],
    )
    biomass_yield, substrate_yield = (
        acetate_flux / biomass_flux,
        acetate_flux / glucose_flux,
    )

    biomass_cell = worksheet.cell(row=current_line + 9, column=2)
    glucose_cell = worksheet.cell(row=current_line + 9, column=3)
    acetate_cell = worksheet.cell(row=current_line + 9, column=4)
    biomass_yield_cell = worksheet.cell(row=current_line + 9, column=5)
    substrate_yield_cell = worksheet.cell(row=current_line + 9, column=6)

    worksheet.cell(row=current_line + 9, column=1).value = ko_target
    biomass_cell.value = biomass_flux
    biomass_cell.fill = _choose_color(biomass_flux, wt_biomass_flux)
    glucose_cell.value = glucose_flux
    glucose_cell.fill = _choose_color(glucose_flux, wt_glucose_flux)
    acetate_cell.value = acetate_flux
    acetate_cell.fill = _choose_color(acetate_flux, wt_acetate_flux)
    biomass_yield_cell.value = biomass_yield
    biomass_yield_cell.fill = _choose_color(biomass_yield, wt_biomass_yield)
    substrate_yield_cell.value = substrate_yield
    substrate_yield_cell.fill = _choose_color(substrate_yield, wt_substrate_yield)

# Save the workbook to an XLSX file
ensure_folder_existence("./examples/iCH360/ko_analysis_spreadsheet")
workbook.save("./examples/iCH360/ko_analysis_spreadsheet/ko_analysis_spreadsheet.xlsx")
